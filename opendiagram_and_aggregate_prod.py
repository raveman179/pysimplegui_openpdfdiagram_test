"""
動作

1.受注No.のバーコードをスキャナで読み込んで、受注No.をフォームに入力する
2.図面を開くボタンを左クリック
3."生産計画シート"のFS受注No.の列から入力された受注Noを検索する。
    受注No.が存在するなら
        最後の列'済'欄にチェックして、
        品目名のセルに入っている値を取得。ー＞変数:prod_nameへ代入する。
    存在しないなら
        "受注No.が存在しません"とかエラーを返す。
4.図面フォルダのパス内から prod_name + '.pdf'と合致するファイルを開く。
"""

import PySimpleGUI as sg
import openpyxl
from os.path import basename
import subprocess
import re

def get_prodname(order_num):
    """
    受注No.検索処理

    xlsx形式の生産計画シートを開いて品名を返す
    
    args: Fs受注No.(品質経歴書バーコード⑨)
    return: 該当受注No.の品目名
    """
    wb = openpyxl.load_workbook('./production_schedule.xlsx')
    sheet = wb.active 

    prod_sheet = wb['Sheet1']
    prod_name = ""

    for row in prod_sheet.iter_rows(min_row=2):
        if order_num == row[0].value:
            row1 = row[1].value
            prod_name = re.sub(r'\s.*ΩJ', '', row1)
            sheet[row[7].coordinate] = "作業済み"
            break

    wb.save('production_schedule.xlsx')

    return prod_name

def get_diagram_num(prod_name):
    """
    図面ファイル検索処理

    変換テーブルから図面のファイル名を取得する

    args: get_prodname()で取得した製品名
    return: 開こうとするPDF図面の絶対パス
    """
    # ------図面フォルダのパス
    diagram_dir_path = "図面の入っているフォルダ名"

    # ------変換テーブルで取得した図面No.
    diagram_filename = ""

    wb = openpyxl.load_workbook('./diagramnum.xlsx')

    sheet = wb['Sheet1']
    for row in sheet.iter_rows(min_row=2):
        if prod_name == row[0].value:
            diagram_filename = row[1].value
    
    pdf_abs_path = diagram_dir_path + diagram_filename

    return pdf_abs_path

sg.theme('Dark Green 1')

layout = [
    [sg.Text('     受注No.(品質経歴書⑨)バーコードを読み込む', pad=((10,10),(20,20))), sg.InputText(key='order_num')],
    [sg.Button('図面を開く', size=(40,1), key='open_diagram'),sg.Button('閉じる', size=(40,1), key='Quit')]
]

window = sg.Window('生産計画、図面チェックプログラム(仮)', layout)

while True:
    event, values = window.read()
    print('イベント:', event ,', 値:',values) # 確認表示
    
    prod_name = get_prodname(values['order_num'])

    # 生産計画シート内に読み込んだFS受注No.が存在しない
    
    if prod_name == '':    
        sg.Popup('受注No.が存在しません')
        #閉じるときに一回開く。意図しない動作なので修正。
        
    pdf_abs_path = get_diagram_num(prod_name)
    
    # 図面のフルパスを投げてpopen関数で図面ファイルを開く
    if event == 'open_diagram':
        subprocess.Popen(['start', pdf_abs_path], shell=True)

    if event == sg.WIN_CLOSED or event == 'Quit':
        break

window.close()
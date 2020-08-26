import openpyxl
from pprint import pprint
import re
# diagram_number = openpyxl.load_workbook('./diagramNo.xlsx')

# sheet = diagram_number['Sheet1']
# cell = sheet['A1:A10']

# pprint(cell.value)
def open_prod_schedule(order_num):
    wb = openpyxl.load_workbook('./production_schedule.xlsx')
    sheet = wb.active 

    prod_sheet = wb['Sheet1']
    prod_name = ""
    fixed = ""

    for row in prod_sheet.iter_rows(min_row=2):
        if row[0].value == order_num:
            row1 = row[1].value
            prod_name = re.sub(r'\s.*ΩJ', '', row1)
            sheet[row[7].coordinate] = "fixed"
            break
    else:
        print('該当する受注No.がありません')

    wb.save('production_schedule.xlsx')

    return prod_name

print(open_prod_schedule('07463620'))


"""
動作

1.受注No.のバーコードをスキャナで読み込んで、受注No.をフォームに入力する
2.図面を開くボタンを左クリック
3."生産計画シート"のFS受注No.の列から入力された受注Noを検索する。
    受注No.が存在するなら
        最後の列'済'欄にチェックして、
        品目名のセルに入っている値を取得。ー＞変数:prod_nameへ代入する。
    存在しないなら
        "受注No.が存在しません"とかエラーを返す。
4.図面フォルダのパス内から prod_name + '.pdf'と合致するファイルを開く。
    変換テーブルを検索して図面が存在するなら
        Popenで図面を開く
    何らかの理由(横型以外の品名等)で図面がフォルダ内に存在しない場合
        "図面ファイルが存在しません"エラーを返す。
"""


import openpyxl
import re

class sheet_search:
    def __init__(self):
        # ------図面フォルダのパス
        self.diagram_dir_path = "C:/Users/製造1課 組込/Desktop/S4_組込(C20)"

        # ------生産計画表のファイル名
        self.production_schedule = './production_schedule.xlsx'

        # ------品名<-->PDF図面ファイル変換テーブルのファイル名
        self.diagram_num = './diagramnum.xlsx'

        # ------変換テーブルで取得した図面No.
        self.diagram_filename = ''


    def get_prodname(self, order_num):
        """
        受注No.検索処理

        xlsx形式の生産計画シートを開いて品名を返す
        
        args: Fs受注No.(品質経歴書バーコード⑨)
        return: 該当受注No.の品目名
        """
        wb = openpyxl.load_workbook(self.production_schedule)
        sheet = wb.active 

        prod_sheet = wb['Sheet1']
        prod_name = ""

        for row in prod_sheet.iter_rows(min_row=2):
            if order_num == row[0].value:
                row1 = row[1].value
                sheet[row[7].coordinate] = "作業済み"
                if row1.endswith('ΩJ'):
                    prod_name = re.sub(r'\s.*ΩJ', '', row1)
                else:
                    prod_name = row1 # そのまま返す。図面ファイル検索時にエラーが出ないことを検証すること。
                break
            else:
                prod_name = "n/a"

        wb.save(self.production_schedule)

        return prod_name
    
    def get_diagram_num(self, prod_name):
        """
        図面ファイル検索処理

        変換テーブルから図面のファイル名を取得する

        args: get_prodname()で取得した製品名
        return: 開こうとするPDF図面の絶対パス
        """
        
        wb = openpyxl.load_workbook(self.diagram_num)

        sheet = wb['Sheet1']
        for row in sheet.iter_rows(min_row=2):
            if prod_name == row[0].value:
                self.diagram_filename = row[1].value
            else:
                self.diagram_filename = "n/a"
                return self.diagram_filename

        pdf_abs_path = self.diagram_dir_path + '/' + self.diagram_filename

        return pdf_abs_path


if __name__ == "__main__":
    sheetsearch = sheet_search()
    
    prod_name = sheetsearch.get_prodname('00')
    print(prod_name)

    abs_path = sheetsearch.get_diagram_num(prod_name)
    # print(abs_path)